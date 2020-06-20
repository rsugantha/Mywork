'''Create student data and modify using openpyxl Grade: 20
Create an Excel file that contains student data (Name, StudentID, Date of Birth) using openpyxl module.
I. The file should contain at least 5 records. Save the file in a desired location.
II. Calculate the age of all the students in the file, based on their DoB
Display the age of all the students'''

def writetoxl(st):
    from openpyxl import Workbook
    wb = Workbook()

    ws = wb.active                   # grab the active worksheet
    ws.title = "Student data"        # Set the title of worksheet

    for i in range(len(st)) :
        ws.append(st[i])             # Write the Student data to work sheet

    wb.save("Student data.xlsx")     # Save the file

def readxl(agear):
    import openpyxl
    path="C:\\Users\\Rishav\\PycharmProjects\\students\\Student data.xlsx"
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    nr = ws.max_row
    nc = ws.max_column
    dobcol = 0
    for i in range(1,nc+1):                     #Find which column in the excel has DOB details
        cl = ws.cell(1,i)
        clv = cl.value
        if clv == 'DOB':
            dobcol = i                          #DOB column

    for i in range(2,nr+1):                     #Create an agelist
            st = ws.cell(i,dobcol)
            stv=st.value.split('-')
            by = stv[0]
            agear = calage(by)                  #Call age calucation fn using birthyear
    return agear

def calage(by):
    import datetime as dt                       #import data time module
    today = dt.datetime.now()
    cy = today.year                             #Current year
    age = int(cy) - int(by)                     #Calulate age by finding the diff between birthyear and Current year
    agear.append(age)
    return agear

#Main
#Student Data
st = [['Name', 'Student ID', 'DOB'], ['ABC', '001', '1989-01-01'], ['ABCD', '002', '1990-02-02'],
      ['ABCDE', '003', '1991-03-02'],
      ['ABCDEF', '004', '1988-04-02'], ['ABCDEFG', '005', '1993-05-02']]

#Call Write function to write data to excel
writetoxl(st)

#Read the Dob column and calculate the age
agear=[]
agelist = readxl(agear)

#Append the age to student data
for i in range(len(st)):
    if i == 0:
        st[i].append('Age')            #Column name -Age
    else:
        st[i].append(agelist[i-1])     #Age

#Print the student data along with age in tabular format
from tabulate import tabulate
print(tabulate(st))